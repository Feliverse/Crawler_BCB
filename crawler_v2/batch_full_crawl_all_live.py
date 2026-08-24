from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
import threading
import unicodedata
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN DEL MAPEO PROFUNDO
# ============================================================

EXPECTED_SOURCES = 52

# Varias fuentes a la vez. Cada crawler mantiene su propia pausa aleatoria.
MAX_WORKERS = 4

# Límites altos pero finitos. Si una fuente llega a uno de estos
# límites, NO se considera completamente mapeada.
MAX_PAGES = 10000
MAX_DEPTH = 20
MAX_FILES = 100000

REQUEST_TIMEOUT = 15
CRAWLER_TIMEOUT = 7200  # 2 horas máximo por fuente

DELAY_SECONDS = 0.30
RANDOM_DELAY_MIN = 0.30
RANDOM_DELAY_MAX = 0.90

API_PAGINATION_MAX_PAGES = 20
MAX_OPENAPI_ENDPOINTS = 100


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent

SOURCES_DIR = SCRIPT_DIR / "sources"
OUTPUT_DIR = SCRIPT_DIR / "output"
FULL_OUTPUT_DIR = OUTPUT_DIR / "full_map"
LOG_DIR = OUTPUT_DIR / "full_logs"

MAIN_PATH = SCRIPT_DIR / "main.py"


# ============================================================
# BLOQUEOS PARA CONFIGS COMPARTIDAS / SALIDA DE TERMINAL
# ============================================================

PRINT_LOCK = threading.Lock()
CONFIG_LOCKS_LOCK = threading.Lock()
CONFIG_LOCKS: dict[str, threading.Lock] = {}


def get_config_lock(source: dict) -> threading.Lock:
    existing = existing_source_config(source)

    if existing is not None:
        key = str(existing.resolve()).lower()
    else:
        key = f"generated:{source['fila_excel']}:{slugify(source['fuente'])}"

    with CONFIG_LOCKS_LOCK:
        lock = CONFIG_LOCKS.get(key)
        if lock is None:
            lock = threading.Lock()
            CONFIG_LOCKS[key] = lock
        return lock


# ============================================================
# CONFIGURACIONES / ADAPTERS YA EXISTENTES
# ============================================================

# OJO: esto NO reutiliza resultados anteriores.
# Solo reutiliza la configuración/adaptador correcto y vuelve
# a ejecutar el crawler desde cero.
KNOWN_SOURCE_CONFIGS = {
    "AE": "aetn",
    "AETN": "aetn",
    "ASFI": "asfi",
    "ASFI - BCB": "bcb",
    "ASOFIN": "asofin",
    "ATT": "att",
    "BBV": "bbv",
    "BCB": "bcb",
    "Data.Gov": "datagov",
    "IBCE - CAO": "ibce",
    "MDRyT": "mdryt",
    "Min. Educación": "minedu",
    "SENAMHI": "senamhi",
    "SIGMA": "sigma",
    "SNIS": "snis",
    "TRANSTATS": "transtats",
}


# ============================================================
# UTILIDADES
# ============================================================

def clean_text(value: object) -> str:
    return " ".join(str(value or "").split()).strip()


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "fuente"


def normalize_url(url: str) -> str:
    value = clean_text(url)
    if not value:
        return ""
    if not value.startswith(("http://", "https://")):
        value = "https://" + value
    return value.rstrip("/")


def domain_from_url(url: str) -> str:
    hostname = (urlparse(url).hostname or "").lower()
    if hostname.startswith("www."):
        hostname = hostname[4:]
    return hostname


# ============================================================
# EXCEL
# ============================================================

def find_excel() -> Path:
    candidates = [
        ROOT_DIR / "resultadosPrimerCrawleo.xlsx",
        ROOT_DIR / "resultadosPrimerCrawleo_actualizado.xlsx",
        Path.cwd() / "resultadosPrimerCrawleo.xlsx",
        Path.cwd() / "resultadosPrimerCrawleo_actualizado.xlsx",
    ]

    for path in candidates:
        if path.exists():
            return path

    raise FileNotFoundError("No encontré resultadosPrimerCrawleo.xlsx.")


def read_sources_from_excel(excel_path: Path) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook.active

    header_row = None
    columns: dict[str, int] = {}

    for row_number in range(1, min(sheet.max_row, 15) + 1):
        values: dict[str, int] = {}

        for column_number in range(1, sheet.max_column + 1):
            value = clean_text(
                sheet.cell(row=row_number, column=column_number).value
            )
            if value:
                values[value.lower()] = column_number

        if "fuente" in values and "dirección web" in values:
            header_row = row_number
            columns = values
            break

    if header_row is None:
        raise RuntimeError("No encontré las columnas Fuente y Dirección web.")

    source_col = columns["fuente"]
    url_col = columns["dirección web"]
    institution_col = columns.get("institución")
    crawler_col = columns.get("crawler")

    sources: list[dict] = []
    skipped: list[dict] = []

    for row_number in range(header_row + 1, sheet.max_row + 1):
        source = clean_text(sheet.cell(row=row_number, column=source_col).value)
        raw_url = clean_text(sheet.cell(row=row_number, column=url_col).value)

        institution = ""
        if institution_col:
            institution = clean_text(
                sheet.cell(row=row_number, column=institution_col).value
            )

        assigned = ""
        if crawler_col:
            assigned = clean_text(
                sheet.cell(row=row_number, column=crawler_col).value
            )

        if not source and not raw_url:
            continue

        if not source or not raw_url:
            skipped.append(
                {
                    "fila_excel": row_number,
                    "fuente": source,
                    "url": raw_url,
                    "motivo": "falta Fuente" if not source else "falta Dirección web",
                }
            )
            continue

        sources.append(
            {
                "fila_excel": row_number,
                "fuente": source,
                "institucion": institution,
                "url": normalize_url(raw_url),
                "crawler_asignado": assigned,
            }
        )

    workbook.close()
    return sources, skipped


# ============================================================
# INSPECCIÓN DEL JSON FINAL
# ============================================================

def collect_resources(node: object, resources: list[dict]) -> None:
    if not isinstance(node, dict):
        return

    if isinstance(node.get("descripcion"), str) and isinstance(
        node.get("url_descarga"), str
    ):
        resources.append(node)
        return

    for value in node.values():
        collect_resources(value, resources)


def inspect_output_json(output_path: Path) -> dict:
    empty = {
        "recursos_json": 0,
        "urls_unicas": 0,
        "duplicados_url": 0,
        "formatos": {},
    }

    if not output_path.exists():
        return empty

    try:
        with output_path.open("r", encoding="utf-8") as file:
            payload = json.load(file)
    except Exception:
        return empty

    resources: list[dict] = []
    collect_resources(payload.get("ESTADISTICAS", {}), resources)

    urls = [
        clean_text(item.get("url_descarga"))
        for item in resources
        if clean_text(item.get("url_descarga"))
    ]

    counter: Counter[str] = Counter()

    for item in resources:
        file_type = clean_text(
            item.get("tipo_archivo", "DESCONOCIDO")
        ).upper() or "DESCONOCIDO"

        if file_type == "API":
            api_format = clean_text(item.get("formato")).upper()
            if api_format:
                counter[f"API:{api_format}"] += 1
                continue

        counter[file_type] += 1

    return {
        "recursos_json": len(resources),
        "urls_unicas": len(set(urls)),
        "duplicados_url": len(urls) - len(set(urls)),
        "formatos": dict(sorted(counter.items())),
    }


# ============================================================
# EXTRAER MÉTRICAS DEL MAIN
# ============================================================

def extract_number(output: str, pattern: str) -> int | None:
    match = re.search(pattern, output, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def extract_float(output: str, pattern: str) -> float | None:
    match = re.search(pattern, output, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def extract_text(output: str, pattern: str) -> str:
    match = re.search(pattern, output, flags=re.IGNORECASE)
    if not match:
        return ""
    return clean_text(match.group(1))


# ============================================================
# CONFIGURACIÓN POR FUENTE
# ============================================================

def existing_source_config(source: dict) -> Path | None:
    configured_id = KNOWN_SOURCE_CONFIGS.get(source["fuente"])

    if configured_id:
        path = SOURCES_DIR / f"{configured_id}.json"
        if path.exists():
            return path

    # Si existe una config cuyo nombre coincide con la fuente,
    # también la usamos.
    for candidate in (
        slugify(source["fuente"]),
        slugify(source["institucion"]),
    ):
        path = SOURCES_DIR / f"{candidate}.json"
        if path.exists():
            return path

    return None


def prepare_source_config(source: dict) -> dict:
    """
    Si existe config real, mantiene su id para no perder el
    adapter específico. Solo eleva temporalmente los límites.
    Luego se restaura exactamente el archivo original.
    """

    existing_path = existing_source_config(source)
    generated = existing_path is None
    original_bytes: bytes | None = None

    if existing_path is not None:
        config_path = existing_path
        original_bytes = config_path.read_bytes()

        with config_path.open("r", encoding="utf-8-sig") as file:
            config = json.load(file)

        source_id = str(config.get("id_fuente", config_path.stem))

    else:
        source_id = (
            f"full_{source['fila_excel']}_"
            f"{slugify(source['fuente'])}"
        )
        config_path = SOURCES_DIR / f"{source_id}.json"

        config = {
            "id_fuente": source_id,
            "nombre": source["institucion"] or source["fuente"],
            "base_url": source["url"],
            "allowed_domains": [domain_from_url(source["url"])],
            "entrypoints": [source["url"]],
            "auto_url_variants": True,
        }

    # Mapeo profundo.
    config["max_pages"] = MAX_PAGES
    config["max_depth"] = MAX_DEPTH
    config["max_files"] = MAX_FILES

    try:
        current_timeout = int(config.get("request_timeout", REQUEST_TIMEOUT))
    except (TypeError, ValueError):
        current_timeout = REQUEST_TIMEOUT

    config["request_timeout"] = max(REQUEST_TIMEOUT, current_timeout)
    config.setdefault("delay_seconds", DELAY_SECONDS)
    config.setdefault("random_delay_min", RANDOM_DELAY_MIN)
    config.setdefault("random_delay_max", RANDOM_DELAY_MAX)
    config.setdefault("auto_url_variants", True)
    config.setdefault("discover_openapi_endpoints", True)
    config.setdefault("crawl_api_documentation", True)
    config.setdefault("max_openapi_endpoints", MAX_OPENAPI_ENDPOINTS)

    if "api_pagination" not in config:
        config["api_pagination"] = {
            "enabled": True,
            "max_pages": API_PAGINATION_MAX_PAGES,
        }

    with config_path.open("w", encoding="utf-8") as file:
        json.dump(config, file, ensure_ascii=False, indent=2)

    return {
        "source_id": source_id,
        "config": config,
        "config_path": config_path,
        "original_bytes": original_bytes,
        "generated": generated,
    }


def restore_source_config(prepared: dict) -> None:
    config_path: Path = prepared["config_path"]
    original_bytes = prepared["original_bytes"]

    if original_bytes is not None:
        config_path.write_bytes(original_bytes)
        return

    if prepared["generated"] and config_path.exists():
        try:
            config_path.unlink()
        except OSError:
            pass


# ============================================================
# CLASIFICAR RESULTADO
# ============================================================

def classify_result(
    *,
    returncode: int | None,
    timed_out: bool,
    stop_reason: str,
    crawler_errors: int,
    resources: int,
    console_output: str,
) -> tuple[str, str]:

    if timed_out:
        return "TIMEOUT", "La fuente excedió el tiempo máximo."

    if returncode is not None and returncode != 0:
        return "ERROR", f"main.py terminó con código {returncode}."

    if resources == 0 and (
        "HTTP=403" in console_output
        or "forbidden" in console_output.lower()
    ):
        return "FUENTE_DENEGADA", "La fuente respondió HTTP 403."

    stop = stop_reason.lower().strip()

    if stop in {"max_pages", "max_files"}:
        return (
            "INCOMPLETO_LIMITE",
            f"El crawler alcanzó {stop_reason}; no se considera mapeo completo.",
        )

    if resources == 0:
        return "SIN_RECURSOS", "Terminó sin recursos catalogados."

    if crawler_errors > 0:
        return (
            "MAPEADO_CON_ERRORES",
            "Generó JSON, pero hubo errores durante la exploración.",
        )

    if stop == "frontier_exhausted":
        return (
            "MAPEADO",
            "La frontera alcanzable quedó agotada sin errores.",
        )

    return "REVISAR", "El motivo de parada requiere revisión."


# ============================================================
# EJECUTAR UNA FUENTE
# ============================================================

def crawl_one(source: dict, position: int, total: int) -> dict:
    source_name = source["fuente"]

    with PRINT_LOCK:
        print()
        print("=" * 80)
        print(f"[{position}/{total}] INICIO {source_name}")
        print("=" * 80)

    # Dos filas pueden reutilizar la misma configuración real (por ejemplo BCB).
    # Solo esas filas se serializan entre sí; las demás corren en paralelo.
    config_lock = get_config_lock(source)

    with config_lock:
        prepared = prepare_source_config(source)

        source_id = prepared["source_id"]
        config = prepared["config"]
        output_id = str(config.get("id_fuente", source_id))

        canonical_output = OUTPUT_DIR / f"{output_id}.json"
        final_output = FULL_OUTPUT_DIR / (
            f"{int(source['fila_excel']):03d}_{slugify(source_name)}.json"
        )
        log_path = LOG_DIR / (
            f"{int(source['fila_excel']):03d}_{slugify(source_name)}.log"
        )

        # Evita confundir una salida previa con la salida de esta corrida.
        try:
            if canonical_output.exists():
                canonical_output.unlink()
        except OSError:
            pass

        try:
            if final_output.exists():
                final_output.unlink()
        except OSError:
            pass

        started = time.monotonic()
        timed_out = False
        returncode: int | None = None

        environment = os.environ.copy()
        environment["PYTHONIOENCODING"] = "utf-8"
        environment["PYTHONUNBUFFERED"] = "1"

        process: subprocess.Popen | None = None
        reader_thread: threading.Thread | None = None

        # Solo mostramos en terminal actividad útil. El log conserva TODO.
        last_progress_print = [0.0]

        def should_echo(line: str) -> bool:
            text = line.strip()
            if not text:
                return False

            important = (
                "DATASETS=",
                "DATASET ACTUALIZADO",
                "API OMITIDA",
                "API PAGINACION",
                "OPENAPI",
                "RESULTADO",
                "Páginas visitadas:",
                "Archivos encontrados:",
                "Datasets web:",
                "Errores:",
                "Motivo de parada:",
                "JSON:",
                "Trazabilidad:",
                "Estado:",
                "Utilizada:",
                "Fallback:",
            )

            if any(token in text for token in important):
                return True

            # Las líneas t=... son el pulso del crawler. Las limitamos a una
            # aproximadamente cada 3 segundos por fuente para no inundar consola.
            if " t=" in text and " | pag=" in text and " | cola=" in text:
                now = time.monotonic()
                if now - last_progress_print[0] >= 3.0:
                    last_progress_print[0] = now
                    return True

            return False

        try:
            with log_path.open("w", encoding="utf-8") as log_file:
                process = subprocess.Popen(
                    [sys.executable, str(MAIN_PATH), source_id],
                    cwd=str(SCRIPT_DIR),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    bufsize=1,
                    env=environment,
                )

                def reader() -> None:
                    assert process is not None
                    assert process.stdout is not None

                    for line in process.stdout:
                        log_file.write(line)
                        log_file.flush()

                        if should_echo(line):
                            with PRINT_LOCK:
                                print(
                                    f"[{source_name}] {line.rstrip()}",
                                    flush=True,
                                )

                reader_thread = threading.Thread(
                    target=reader,
                    name=f"reader-{slugify(source_name)}",
                    daemon=True,
                )
                reader_thread.start()

                try:
                    returncode = process.wait(timeout=CRAWLER_TIMEOUT)

                except subprocess.TimeoutExpired:
                    timed_out = True

                    with PRINT_LOCK:
                        print(
                            f"[{source_name}] TIMEOUT después de "
                            f"{CRAWLER_TIMEOUT}s. Finalizando proceso...",
                            flush=True,
                        )

                    process.kill()
                    returncode = process.wait(timeout=30)
                    log_file.write("\n[BATCH] TIMEOUT DE FUENTE\n")
                    log_file.flush()

                finally:
                    if reader_thread is not None:
                        reader_thread.join(timeout=10)

        finally:
            restore_source_config(prepared)

        total_duration = time.monotonic() - started

        try:
            console_output = log_path.read_text(
                encoding="utf-8",
                errors="replace",
            )
        except Exception:
            console_output = ""

        pages = extract_number(
            console_output,
            r"Páginas visitadas:\s*(\d+)",
        ) or 0

        files = extract_number(
            console_output,
            r"Archivos encontrados:\s*(\d+)",
        ) or 0

        datasets = extract_number(
            console_output,
            r"Datasets web:\s*(\d+)",
        ) or 0

        errors = extract_number(
            console_output,
            r"Errores:\s*(\d+)",
        ) or 0

        crawler_duration = extract_float(
            console_output,
            r"Duración:\s*([\d.]+)",
        ) or 0.0

        stop_reason = extract_text(
            console_output,
            r"Motivo de parada:\s*([^\r\n]+)",
        )

        used_url = (
            extract_text(console_output, r"Utilizada:\s*([^\r\n]+)")
            or extract_text(console_output, r"URL de trabajo:\s*([^\r\n]+)")
        )

        fallback_text = extract_text(
            console_output,
            r"Fallback:\s*([^\r\n]+)",
        ).lower()

        used_fallback = fallback_text in {"sí", "si", "yes", "true"}

        # main.py hace replace atómico al terminar; si existe aquí pertenece
        # a esta ejecución porque lo eliminamos antes de arrancar.
        if canonical_output.exists():
            try:
                shutil.copy2(canonical_output, final_output)
            except OSError:
                pass

        metrics = inspect_output_json(final_output)

        status, detail = classify_result(
            returncode=returncode,
            timed_out=timed_out,
            stop_reason=stop_reason,
            crawler_errors=errors,
            resources=metrics["recursos_json"],
            console_output=console_output,
        )

        formats_text = " | ".join(
            f"{key}={value}"
            for key, value in metrics["formatos"].items()
        )

        with PRINT_LOCK:
            print()
            print(f"[{position}/{total}] FIN {source_name}")
            print(f"  Estado:        {status}")
            print(f"  Páginas:       {pages}")
            print(f"  Archivos:      {files}")
            print(f"  Datasets:      {datasets}")
            print(f"  Recursos JSON: {metrics['recursos_json']}")
            print(f"  Errores:       {errors}")
            print(f"  Parada:        {stop_reason or '-'}")
            print(f"  Duración:      {total_duration:.2f}s")

            if formats_text:
                print(f"  Formatos:      {formats_text}")

            print(
                f"  JSON:          "
                f"{final_output if final_output.exists() else 'NO GENERADO'}"
            )
            print(f"  Log:           {log_path}")

        return {
            **source,
            "id_fuente": output_id,
            "url_utilizada": used_url,
            "uso_fallback": used_fallback,
            "paginas": pages,
            "archivos": files,
            "datasets": datasets,
            "recursos_json": metrics["recursos_json"],
            "urls_unicas": metrics["urls_unicas"],
            "duplicados_url": metrics["duplicados_url"],
            "formatos": metrics["formatos"],
            "errores": errors,
            "motivo_parada": stop_reason,
            "duracion_crawler": crawler_duration,
            "duracion_total": round(total_duration, 2),
            "returncode": returncode,
            "resultado": status,
            "detalle": detail,
            "json_output": str(final_output) if final_output.exists() else "",
            "log_output": str(log_path),
        }


# ============================================================
# GUARDAR RESUMEN
# ============================================================

def save_results(results: list[dict], skipped: list[dict]) -> None:
    json_path = OUTPUT_DIR / "mapeo_completo_resultados.json"
    csv_path = OUTPUT_DIR / "mapeo_completo_resultados.csv"
    skipped_path = OUTPUT_DIR / "mapeo_completo_omitidos.json"

    with json_path.open("w", encoding="utf-8") as file:
        json.dump(results, file, ensure_ascii=False, indent=2)

    with skipped_path.open("w", encoding="utf-8") as file:
        json.dump(skipped, file, ensure_ascii=False, indent=2)

    fieldnames = [
        "fila_excel",
        "fuente",
        "institucion",
        "url",
        "url_utilizada",
        "uso_fallback",
        "crawler_asignado",
        "id_fuente",
        "paginas",
        "archivos",
        "datasets",
        "recursos_json",
        "urls_unicas",
        "duplicados_url",
        "formatos",
        "errores",
        "motivo_parada",
        "duracion_crawler",
        "duracion_total",
        "returncode",
        "resultado",
        "detalle",
        "json_output",
        "log_output",
    ]

    with csv_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
            delimiter=";",
        )
        writer.writeheader()

        for item in results:
            row = dict(item)
            row["formatos"] = json.dumps(
                row.get("formatos", {}),
                ensure_ascii=False,
                sort_keys=True,
            )
            writer.writerow(row)


# ============================================================
# RESUMEN FINAL
# ============================================================

def print_final_summary(results: list[dict], skipped: list[dict]) -> None:
    statuses = Counter(item["resultado"] for item in results)
    formats: Counter[str] = Counter()

    for item in results:
        formats.update(item.get("formatos", {}))

    total_resources = sum(
        int(item.get("recursos_json", 0))
        for item in results
    )

    print()
    print("=" * 80)
    print("RESUMEN FINAL - MAPEO PROFUNDO")
    print("=" * 80)

    print(f"Fuentes procesadas: {len(results)}")
    print(f"Filas omitidas:     {len(skipped)}")
    print(f"Recursos JSON:      {total_resources}")

    print()
    print("Estados:")

    for status, total in sorted(statuses.items()):
        print(f"  {status:<24} {total}")

    if formats:
        print()
        print("Formatos encontrados:")

        for file_type, total in sorted(formats.items()):
            print(f"  {file_type:<24} {total}")

    if skipped:
        print()
        print("FILAS OMITIDAS DEL EXCEL:")

        for item in skipped:
            print(
                f"  fila={item['fila_excel']} | "
                f"fuente={item['fuente'] or '-'} | "
                f"{item['motivo']}"
            )

    print()
    print(
        "MAPEADO = terminó por frontier_exhausted y sin errores."
    )
    print(
        "Cualquier otro estado requiere corrección o una segunda pasada."
    )


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    SOURCES_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    FULL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    excel_path = find_excel()
    sources, skipped = read_sources_from_excel(excel_path)

    print("=" * 80)
    print("BATCH - MAPEO PROFUNDO DE TODAS LAS FUENTES")
    print("=" * 80)
    print(f"Excel: {excel_path}")
    print(f"Fuentes válidas encontradas: {len(sources)}")

    if len(sources) != EXPECTED_SOURCES:
        print()
        print(
            f"ADVERTENCIA: esperábamos {EXPECTED_SOURCES} fuentes, "
            f"pero el Excel tiene {len(sources)} filas válidas."
        )

    print()
    print("Límites de seguridad:")
    print(f"  max_pages = {MAX_PAGES}")
    print(f"  max_depth = {MAX_DEPTH}")
    print(f"  max_files = {MAX_FILES}")
    print(
        "Una fuente solo será MAPEADO si termina por frontier_exhausted."
    )

    print(f"  workers = {MAX_WORKERS}")
    print("  salida = EN VIVO (además se guarda log completo por fuente)")

    results: list[dict] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {
            executor.submit(
                crawl_one,
                source,
                position,
                len(sources),
            ): source
            for position, source in enumerate(sources, start=1)
        }

        for future in as_completed(future_map):
            source = future_map[future]

            try:
                result = future.result()
            except Exception as error:
                result = {
                    **source,
                    "id_fuente": "",
                    "url_utilizada": "",
                    "uso_fallback": False,
                    "paginas": 0,
                    "archivos": 0,
                    "datasets": 0,
                    "recursos_json": 0,
                    "urls_unicas": 0,
                    "duplicados_url": 0,
                    "formatos": {},
                    "errores": 1,
                    "motivo_parada": "batch_exception",
                    "duracion_crawler": 0.0,
                    "duracion_total": 0.0,
                    "returncode": None,
                    "resultado": "ERROR",
                    "detalle": f"{type(error).__name__}: {error}",
                    "json_output": "",
                    "log_output": "",
                }

                with PRINT_LOCK:
                    print(
                        f"[{source['fuente']}] ERROR BATCH | "
                        f"{type(error).__name__}: {error}",
                        flush=True,
                    )

            results.append(result)

            # Guardado incremental: si la ejecución se interrumpe, no se
            # pierde lo que ya terminó.
            save_results(
                sorted(results, key=lambda item: int(item["fila_excel"])),
                skipped,
            )

    results.sort(key=lambda item: int(item["fila_excel"]))
    save_results(results, skipped)
    print_final_summary(results, skipped)

    print()
    print(f"JSON individuales: {FULL_OUTPUT_DIR}")
    print(f"Logs individuales: {LOG_DIR}")
    print(
        f"Resumen JSON: "
        f"{OUTPUT_DIR / 'mapeo_completo_resultados.json'}"
    )
    print(
        f"Resumen CSV : "
        f"{OUTPUT_DIR / 'mapeo_completo_resultados.csv'}"
    )


if __name__ == "__main__":
    main()