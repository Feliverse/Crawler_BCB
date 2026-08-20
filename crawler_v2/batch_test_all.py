from __future__ import annotations

import csv
import json
import re
import subprocess
import sys
import unicodedata

from concurrent.futures import (
    ThreadPoolExecutor,
    as_completed,
)
from pathlib import Path
from urllib.parse import urlparse
from urllib.robotparser import RobotFileParser

import requests
from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN RÁPIDA
# ============================================================

MAX_WORKERS = 6

# Primera pasada rápida.
MAX_PAGES = 25
MAX_DEPTH = 3
MAX_FILES = 2000

REQUEST_TIMEOUT = 6
CRAWLER_TIMEOUT = 180
DELAY_SECONDS = 0.10


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent

SOURCES_DIR = SCRIPT_DIR / "sources"
OUTPUT_DIR = SCRIPT_DIR / "output"
MAIN_PATH = SCRIPT_DIR / "main.py"


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)


# ============================================================
# FUENTES QUE YA PROBAMOS Y VALIDAMOS
# ============================================================

PREVALIDATED = {
    "AE": ("aetn", "exitoso"),
    "ASFI - BCB": ("bcb", "exitoso"),
    "ASOFIN": ("asofin", "exitoso"),
    "ATT": ("att", "exitoso"),
    "BBV": ("bbv", "exitoso"),
    "BCB": ("bcb", "exitoso"),
    "Data.Gov": ("datagov", "exitoso"),
    "IBCE - CAO": ("ibce", "exitoso"),
    "MDRyT": ("mdryt", "exitoso"),
    "Min. Educación": ("minedu", "exitoso"),
    "SENAMHI": ("senamhi", "exitoso"),
    "SIGMA": ("sigma", "sin enlace directo"),
    "SNIS": ("snis", "exitoso"),
    "TRANSTATS": ("transtats", "exitoso"),
}


# ============================================================
# UTILIDADES
# ============================================================

def find_excel() -> Path:
    candidates = [
        ROOT_DIR / "resultadosPrimerCrawleo.xlsx",
        ROOT_DIR / "resultadosPrimerCrawleo_actualizado.xlsx",
        Path.cwd() / "resultadosPrimerCrawleo.xlsx",
        Path.cwd() / "resultadosPrimerCrawleo_actualizado.xlsx",
    ]

    for path in candidates:
        if path.exists():
            return path

    raise FileNotFoundError(
        "No encontré resultadosPrimerCrawleo.xlsx. "
        "Colócalo en la raíz del proyecto."
    )


def clean_text(value: object) -> str:
    return " ".join(
        str(value or "").split()
    ).strip()


def slugify(value: str) -> str:
    text = unicodedata.normalize(
        "NFKD",
        value,
    )

    text = "".join(
        char
        for char in text
        if not unicodedata.combining(char)
    )

    text = text.lower()

    text = re.sub(
        r"[^a-z0-9]+",
        "_",
        text,
    )

    text = text.strip("_")

    return text or "fuente"


def normalize_url(url: str) -> str:
    url = clean_text(url)

    if not url:
        return ""

    if not url.startswith(
        ("http://", "https://")
    ):
        url = "https://" + url

    return url.rstrip("/")


def domain_from_url(url: str) -> str:
    hostname = (
        urlparse(url).hostname
        or ""
    ).lower()

    if hostname.startswith("www."):
        hostname = hostname[4:]

    return hostname


# ============================================================
# LEER EXCEL
# ============================================================

def read_sources_from_excel(
    excel_path: Path,
) -> list[dict]:

    workbook = load_workbook(
        excel_path,
        data_only=True,
    )

    sheet = workbook.active

    header_row = None
    columns = {}

    for row_number in range(
        1,
        min(sheet.max_row, 15) + 1,
    ):

        row_values = {}

        for column_number in range(
            1,
            sheet.max_column + 1,
        ):
            value = clean_text(
                sheet.cell(
                    row=row_number,
                    column=column_number,
                ).value
            )

            if value:
                row_values[value.lower()] = (
                    column_number
                )

        if (
            "fuente" in row_values
            and "dirección web" in row_values
        ):
            header_row = row_number
            columns = row_values
            break

    if header_row is None:
        raise RuntimeError(
            "No encontré las columnas Fuente y Dirección web."
        )

    source_col = columns["fuente"]
    url_col = columns["dirección web"]

    institution_col = columns.get(
        "institución"
    )

    crawler_col = columns.get(
        "crawler"
    )

    sources = []

    for row_number in range(
        header_row + 1,
        sheet.max_row + 1,
    ):

        source = clean_text(
            sheet.cell(
                row=row_number,
                column=source_col,
            ).value
        )

        url = normalize_url(
            sheet.cell(
                row=row_number,
                column=url_col,
            ).value
        )

        if not source or not url:
            continue

        institution = ""

        if institution_col:
            institution = clean_text(
                sheet.cell(
                    row=row_number,
                    column=institution_col,
                ).value
            )

        assigned = ""

        if crawler_col:
            assigned = clean_text(
                sheet.cell(
                    row=row_number,
                    column=crawler_col,
                ).value
            )

        sources.append(
            {
                "fila_excel": row_number,
                "fuente": source,
                "institucion": institution,
                "url": url,
                "crawler_asignado": assigned,
            }
        )

    return sources


# ============================================================
# CONTAR DOCUMENTOS JSON
# ============================================================

def count_documents(
    node: object,
) -> int:

    if not isinstance(
        node,
        dict,
    ):
        return 0

    if (
        isinstance(
            node.get("descripcion"),
            str,
        )
        and isinstance(
            node.get("url_descarga"),
            str,
        )
    ):
        return 1

    return sum(
        count_documents(value)
        for value in node.values()
    )


def count_output_documents(
    output_path: Path,
) -> int:

    if not output_path.exists():
        return 0

    try:
        with output_path.open(
            "r",
            encoding="utf-8",
        ) as file:
            data = json.load(file)

        statistics = data.get(
            "ESTADISTICAS",
            {},
        )

        return count_documents(
            statistics
        )

    except Exception:
        return 0


# ============================================================
# PRECHECK HTTP
# ============================================================

def http_precheck(
    url: str,
) -> dict:

    session = requests.Session()

    session.headers.update(
        {
            "User-Agent": USER_AGENT,
        }
    )

    try:
        response = session.get(
            url,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
        )

    except requests.RequestException as error:
        return {
            "ok": False,
            "http": None,
            "final_url": url,
            "robots": "no comprobado",
            "resultado": "REVISAR",
            "detalle": (
                f"Error HTTP: {error}"
            ),
        }

    status = response.status_code
    final_url = response.url

    if status == 403:
        return {
            "ok": False,
            "http": status,
            "final_url": final_url,
            "robots": "no comprobado",
            "resultado": "fuente denegada",
            "detalle": "HTTP 403 Forbidden",
        }

    if status >= 400:
        return {
            "ok": False,
            "http": status,
            "final_url": final_url,
            "robots": "no comprobado",
            "resultado": "REVISAR",
            "detalle": (
                f"HTTP {status}"
            ),
        }

    parsed = urlparse(
        final_url
    )

    robots_url = (
        f"{parsed.scheme}://"
        f"{parsed.netloc}/robots.txt"
    )

    robots_state = "sin bloqueo"

    try:
        robots_response = session.get(
            robots_url,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
        )

        if robots_response.status_code == 200:

            parser = RobotFileParser()

            parser.set_url(
                robots_url
            )

            parser.parse(
                robots_response.text.splitlines()
            )

            allowed = parser.can_fetch(
                "*",
                final_url,
            )

            if not allowed:
                return {
                    "ok": False,
                    "http": status,
                    "final_url": final_url,
                    "robots": "bloqueado",
                    "resultado": (
                        "bloqueo por robot.txt"
                    ),
                    "detalle": (
                        "robots.txt impide "
                        "el acceso al punto de entrada"
                    ),
                }

            robots_state = "permitido"

    except requests.RequestException:
        robots_state = (
            "no comprobado"
        )

    return {
        "ok": True,
        "http": status,
        "final_url": final_url,
        "robots": robots_state,
        "resultado": None,
        "detalle": "",
    }


# ============================================================
# CONFIG TEMPORAL PARA GENERIC ADAPTER
# ============================================================

def create_temporary_source(
    *,
    source_name: str,
    institution: str,
    url: str,
) -> tuple[str, Path]:

    source_id = (
        "batch_"
        + slugify(source_name)
    )

    source_path = (
        SOURCES_DIR
        / f"{source_id}.json"
    )

    domain = domain_from_url(
        url
    )

    config = {
        "id_fuente": source_id,
        "nombre": (
            institution
            or source_name
        ),
        "base_url": url,
        "allowed_domains": [
            domain
        ],
        "entrypoints": [
            url
        ],
        "max_depth": MAX_DEPTH,
        "max_pages": MAX_PAGES,
        "max_files": MAX_FILES,
        "request_timeout": (
            REQUEST_TIMEOUT
        ),
        "delay_seconds": (
            DELAY_SECONDS
        ),
    }

    with source_path.open(
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            config,
            file,
            ensure_ascii=False,
            indent=2,
        )

    return (
        source_id,
        source_path,
    )


# ============================================================
# EXTRAER MÉTRICAS DE CONSOLA
# ============================================================

def extract_number(
    output: str,
    pattern: str,
) -> int | None:

    match = re.search(
        pattern,
        output,
        flags=re.IGNORECASE,
    )

    if not match:
        return None

    try:
        return int(
            match.group(1)
        )
    except ValueError:
        return None


def extract_float(
    output: str,
    pattern: str,
) -> float | None:

    match = re.search(
        pattern,
        output,
        flags=re.IGNORECASE,
    )

    if not match:
        return None

    try:
        return float(
            match.group(1)
        )
    except ValueError:
        return None


def extract_text(
    output: str,
    pattern: str,
) -> str:

    match = re.search(
        pattern,
        output,
        flags=re.IGNORECASE,
    )

    if not match:
        return ""

    return clean_text(
        match.group(1)
    )


# ============================================================
# EJECUTAR UNA FUENTE
# ============================================================

def crawl_one(
    source: dict,
) -> dict:

    source_name = source["fuente"]
    original_url = source["url"]

    print(
        f"[INICIO] {source_name}"
    )

    precheck = http_precheck(
        original_url
    )

    if not precheck["ok"]:
        return {
            **source,
            "http": precheck["http"],
            "robots": precheck["robots"],
            "paginas": 0,
            "archivos": 0,
            "datasets": 0,
            "documentos": 0,
            "errores": 0,
            "motivo_parada": "",
            "duracion": 0,
            "resultado": (
                precheck["resultado"]
            ),
            "detalle": (
                precheck["detalle"]
            ),
        }

    final_url = precheck[
        "final_url"
    ]

    source_id = ""
    source_path = None

    try:
        (
            source_id,
            source_path,
        ) = create_temporary_source(
            source_name=source_name,
            institution=source[
                "institucion"
            ],
            url=final_url,
        )

        output_path = (
            OUTPUT_DIR
            / f"{source_id}.json"
        )

        if output_path.exists():
            output_path.unlink()

        try:
            process = subprocess.run(
                [
                    sys.executable,
                    str(MAIN_PATH),
                    source_id,
                ],
                cwd=str(ROOT_DIR),
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=CRAWLER_TIMEOUT,
            )

        except subprocess.TimeoutExpired:
            return {
                **source,
                "http": precheck["http"],
                "robots": precheck[
                    "robots"
                ],
                "paginas": 0,
                "archivos": 0,
                "datasets": 0,
                "documentos": 0,
                "errores": 0,
                "motivo_parada": (
                    "timeout_batch"
                ),
                "duracion": (
                    CRAWLER_TIMEOUT
                ),
                "resultado": "REVISAR",
                "detalle": (
                    "Crawler excedió "
                    f"{CRAWLER_TIMEOUT}s"
                ),
            }

        console_output = (
            process.stdout
            + "\n"
            + process.stderr
        )

        paginas = extract_number(
            console_output,
            r"Páginas visitadas:\s*(\d+)",
        )

        archivos = extract_number(
            console_output,
            r"Archivos encontrados:\s*(\d+)",
        )

        datasets = extract_number(
            console_output,
            r"Datasets web:\s*(\d+)",
        )

        errores = extract_number(
            console_output,
            r"Errores:\s*(\d+)",
        )

        duracion = extract_float(
            console_output,
            r"Duración:\s*([\d.]+)",
        )

        motivo = extract_text(
            console_output,
            r"Motivo de parada:\s*([^\r\n]+)",
        )

        documentos = (
            count_output_documents(
                output_path
            )
        )

        if documentos > 0:
            resultado = "exitoso"
            detalle = (
                "Generó contratos JSON"
            )

        elif process.returncode == 0:
            resultado = (
                "sin enlace directo"
            )
            detalle = (
                "HTTP accesible pero "
                "no se detectaron recursos"
            )

        else:
            resultado = "REVISAR"
            detalle = (
                "El crawler terminó "
                f"con código {process.returncode}"
            )

        return {
            **source,
            "http": precheck["http"],
            "robots": precheck[
                "robots"
            ],
            "paginas": paginas or 0,
            "archivos": archivos or 0,
            "datasets": datasets or 0,
            "documentos": documentos,
            "errores": errores or 0,
            "motivo_parada": motivo,
            "duracion": (
                duracion or 0
            ),
            "resultado": resultado,
            "detalle": detalle,
        }

    except Exception as error:
        return {
            **source,
            "http": precheck["http"],
            "robots": precheck[
                "robots"
            ],
            "paginas": 0,
            "archivos": 0,
            "datasets": 0,
            "documentos": 0,
            "errores": 0,
            "motivo_parada": "",
            "duracion": 0,
            "resultado": "REVISAR",
            "detalle": str(error),
        }

    finally:
        if (
            source_path
            and source_path.exists()
        ):
            try:
                source_path.unlink()
            except OSError:
                pass


# ============================================================
# RESULTADOS YA VALIDADOS
# ============================================================

def build_prevalidated_result(
    source: dict,
    source_id: str,
    status: str,
) -> dict:

    output_path = (
        OUTPUT_DIR
        / f"{source_id}.json"
    )

    documents = (
        count_output_documents(
            output_path
        )
    )

    return {
        **source,
        "http": 200,
        "robots": "validado previamente",
        "paginas": "",
        "archivos": "",
        "datasets": "",
        "documentos": documents,
        "errores": "",
        "motivo_parada": (
            "validado previamente"
        ),
        "duracion": "",
        "resultado": status,
        "detalle": (
            "Resultado reutilizado "
            "de pruebas anteriores"
        ),
    }


# ============================================================
# GUARDAR RESULTADOS
# ============================================================

def save_results(
    results: list[dict],
) -> None:

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    json_path = (
        OUTPUT_DIR
        / "batch_resultados.json"
    )

    csv_path = (
        OUTPUT_DIR
        / "batch_resultados.csv"
    )

    with json_path.open(
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            results,
            file,
            ensure_ascii=False,
            indent=2,
        )

    fieldnames = [
        "fila_excel",
        "fuente",
        "institucion",
        "url",
        "crawler_asignado",
        "http",
        "robots",
        "paginas",
        "archivos",
        "datasets",
        "documentos",
        "errores",
        "motivo_parada",
        "duracion",
        "resultado",
        "detalle",
    ]

    with csv_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as file:

        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
            delimiter=";",
        )

        writer.writeheader()

        writer.writerows(
            results
        )

    print()
    print(
        f"JSON: {json_path}"
    )

    print(
        f"CSV : {csv_path}"
    )


# ============================================================
# MAIN
# ============================================================

def main() -> None:

    excel_path = find_excel()

    print(
        f"Excel: {excel_path}"
    )

    sources = read_sources_from_excel(
        excel_path
    )

    print(
        f"Fuentes encontradas: "
        f"{len(sources)}"
    )

    results = []

    pending_by_url: dict[
        str,
        list[dict],
    ] = {}

    # --------------------------------------------------------
    # 1. Reutilizar lo que ya probamos.
    # --------------------------------------------------------

    for source in sources:

        prevalidated = PREVALIDATED.get(
            source["fuente"]
        )

        if prevalidated:

            source_id, status = (
                prevalidated
            )

            result = (
                build_prevalidated_result(
                    source,
                    source_id,
                    status,
                )
            )

            results.append(
                result
            )

            print(
                f"[REUTILIZADO] "
                f"{source['fuente']} "
                f"-> {status}"
            )

            continue

        # ----------------------------------------------------
        # URL idéntica = una sola prueba rápida.
        # ----------------------------------------------------

        normalized = normalize_url(
            source["url"]
        )

        pending_by_url.setdefault(
            normalized,
            [],
        ).append(source)

    # --------------------------------------------------------
    # 2. Ejecutar URLs únicas en paralelo.
    # --------------------------------------------------------

    canonical_sources = [
        group[0]
        for group in pending_by_url.values()
    ]

    print()
    print(
        f"URLs nuevas únicas a probar: "
        f"{len(canonical_sources)}"
    )

    print(
        f"Paralelismo: {MAX_WORKERS}"
    )

    print(
        f"Máximo páginas por sitio: "
        f"{MAX_PAGES}"
    )

    print()

    with ThreadPoolExecutor(
        max_workers=MAX_WORKERS,
    ) as executor:

        future_map = {
            executor.submit(
                crawl_one,
                source,
            ): source
            for source in canonical_sources
        }

        for future in as_completed(
            future_map
        ):

            canonical = (
                future_map[future]
            )

            try:
                base_result = (
                    future.result()
                )

            except Exception as error:
                base_result = {
                    **canonical,
                    "http": "",
                    "robots": "",
                    "paginas": 0,
                    "archivos": 0,
                    "datasets": 0,
                    "documentos": 0,
                    "errores": 0,
                    "motivo_parada": "",
                    "duracion": 0,
                    "resultado": "REVISAR",
                    "detalle": str(error),
                }

            same_url_sources = (
                pending_by_url[
                    normalize_url(
                        canonical["url"]
                    )
                ]
            )

            for source in same_url_sources:

                copied = dict(
                    base_result
                )

                copied.update(
                    {
                        "fila_excel": (
                            source[
                                "fila_excel"
                            ]
                        ),
                        "fuente": (
                            source["fuente"]
                        ),
                        "institucion": (
                            source[
                                "institucion"
                            ]
                        ),
                        "url": (
                            source["url"]
                        ),
                        "crawler_asignado": (
                            source[
                                "crawler_asignado"
                            ]
                        ),
                    }
                )

                results.append(
                    copied
                )

            print(
                f"[{base_result['resultado']}] "
                f"{canonical['fuente']} "
                f"| HTTP={base_result['http']} "
                f"| docs="
                f"{base_result['documentos']}"
            )

    # --------------------------------------------------------
    # Orden original del Excel
    # --------------------------------------------------------

    results.sort(
        key=lambda item: (
            item["fila_excel"]
        )
    )

    save_results(
        results
    )

    # --------------------------------------------------------
    # Resumen
    # --------------------------------------------------------

    counts = {}

    for result in results:
        status = result[
            "resultado"
        ]

        counts[status] = (
            counts.get(status, 0)
            + 1
        )

    print()
    print("=" * 72)
    print("RESUMEN FINAL")
    print("=" * 72)

    for status, total in sorted(
        counts.items()
    ):
        print(
            f"{status:<25} {total}"
        )

    print()
    print(
        "IMPORTANTE: los casos 'sin enlace directo' "
        "y 'REVISAR' son los únicos que conviene "
        "analizar después con mayor profundidad."
    )


if __name__ == "__main__":
    main()